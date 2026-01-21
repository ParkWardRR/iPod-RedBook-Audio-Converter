"""XLSX generation with atomic writes."""

import os
import shutil
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from yangon.models.album import Album
from yangon.models.plan import Action
from yangon.models.status import ArtStatus, TagStatus
from yangon.planner.defaults import compute_default_action
from yangon.xlsx.reader import read_xlsx
from yangon.xlsx.schemas import (
    AAC_BITRATE_OPTIONS,
    ACTION_OPTIONS,
    ALBUMS_COLUMNS,
    COLUMN_STYLE_MAP,
    COLUMN_STYLES,
    SCHEMA_VERSION,
    STATUS_VALUES,
    USER_COLUMNS,
)


class XLSXLockError(Exception):
    """XLSX file is locked."""

    pass


def check_xlsx_lock(xlsx_path: Path) -> None:
    """
    Check if XLSX file is locked by another process.

    Raises:
        XLSXLockError: If lock file exists
    """
    # Check for Excel lock files
    lock_patterns = [
        xlsx_path.parent / f".~lock.{xlsx_path.name}#",  # LibreOffice
        xlsx_path.parent / f"~${xlsx_path.name}",  # Excel Windows
    ]

    for lock_path in lock_patterns:
        if lock_path.exists():
            raise XLSXLockError(
                f"XLSX appears to be locked: {lock_path.name}\n"
                "Please close the file in Excel/LibreOffice and try again."
            )


def create_backup(xlsx_path: Path) -> Path | None:
    """
    Create timestamped backup of existing XLSX.

    Returns backup path or None if no backup created.
    """
    if not xlsx_path.exists():
        return None

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_name = f"{xlsx_path.stem}.{timestamp}{xlsx_path.suffix}"
    backup_path = xlsx_path.parent / backup_name

    shutil.copy2(xlsx_path, backup_path)
    return backup_path


def write_xlsx(
    albums: list[Album],
    xlsx_path: Path,
    library_root: Path,
    preserve_user_edits: bool = True,
) -> Path:
    """
    Write albums to XLSX with atomic write and user edit preservation.

    Args:
        albums: List of albums to write
        xlsx_path: Output path for XLSX
        library_root: Root path of music library
        preserve_user_edits: If True, preserve user columns from existing XLSX

    Returns:
        Path to written XLSX

    Raises:
        XLSXLockError: If file is locked
    """
    xlsx_path = Path(xlsx_path).resolve()

    # Check for lock
    check_xlsx_lock(xlsx_path)

    # Load existing user edits if preserving
    existing_data = {}
    if preserve_user_edits and xlsx_path.exists():
        existing_data = read_xlsx(xlsx_path)

    # Create workbook
    wb = Workbook()

    # Create Summary sheet
    _write_summary_sheet(wb, albums, library_root)

    # Create Albums sheet
    _write_albums_sheet(wb, albums, library_root, existing_data)

    # Create Reference sheet (enum options, column legend)
    _write_reference_sheet(wb)

    # Remove default empty sheet if it exists
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # Write atomically
    temp_path = xlsx_path.parent / f"{xlsx_path.name}.tmp"
    try:
        # Create backup of existing file
        if xlsx_path.exists():
            create_backup(xlsx_path)

        # Write to temp file
        wb.save(temp_path)

        # Atomic rename
        os.replace(temp_path, xlsx_path)
    finally:
        # Clean up temp file if it exists
        if temp_path.exists():
            temp_path.unlink()

    return xlsx_path


def _write_summary_sheet(wb: Workbook, albums: list[Album], library_root: Path) -> None:
    """Write Summary tab with rollup statistics."""
    ws = wb.create_sheet("Summary", 0)

    # Header style
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    # Summary data
    now = datetime.now().isoformat()

    # Count statuses
    tag_counts = {status: 0 for status in TagStatus}
    art_counts = {status: 0 for status in ArtStatus}
    action_counts = {action: 0 for action in Action}

    for album in albums:
        tag_counts[album.tag_status] += 1
        art_counts[album.art_status] += 1
        default_action = compute_default_action(album)
        action_counts[default_action] += 1

    summary_rows = [
        ("schema_version", SCHEMA_VERSION),
        ("library_root", str(library_root)),
        ("updated_at", now),
        ("", ""),
        ("Albums Statistics", ""),
        ("total_albums", len(albums)),
        ("total_tracks", sum(album.track_count for album in albums)),
        ("", ""),
        ("Tag Status", ""),
        ("tag_green", tag_counts[TagStatus.GREEN]),
        ("tag_yellow", tag_counts[TagStatus.YELLOW]),
        ("tag_red", tag_counts[TagStatus.RED]),
        ("", ""),
        ("Art Status", ""),
        ("art_green", art_counts[ArtStatus.GREEN]),
        ("art_yellow", art_counts[ArtStatus.YELLOW]),
        ("art_red", art_counts[ArtStatus.RED]),
        ("", ""),
        ("Default Actions", ""),
        ("alac_preserve", action_counts[Action.ALAC_PRESERVE]),
        ("aac", action_counts[Action.AAC]),
        ("pass_mp3", action_counts[Action.PASS_MP3]),
    ]

    for row_idx, (key, value) in enumerate(summary_rows, start=1):
        ws.cell(row=row_idx, column=1, value=key)
        ws.cell(row=row_idx, column=2, value=value)

        # Style headers
        if key and value == "":
            ws.cell(row=row_idx, column=1).font = header_font
            ws.cell(row=row_idx, column=1).fill = header_fill

    # Set column widths
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 50


def _write_albums_sheet(
    wb: Workbook,
    albums: list[Album],
    library_root: Path,
    existing_data: dict,
) -> None:
    """Write Albums tab with all album data."""
    ws = wb.create_sheet("Albums", 1)

    # Styles
    header_font = Font(bold=True)
    header_border = Border(bottom=Side(style="thin"))

    # Column category fills for headers
    editable_fill = PatternFill(
        start_color=COLUMN_STYLES["editable"]["fill_color"],
        end_color=COLUMN_STYLES["editable"]["fill_color"],
        fill_type="solid",
    )
    computed_fill = PatternFill(
        start_color=COLUMN_STYLES["computed"]["fill_color"],
        end_color=COLUMN_STYLES["computed"]["fill_color"],
        fill_type="solid",
    )
    info_fill = PatternFill(
        start_color=COLUMN_STYLES["info"]["fill_color"],
        end_color=COLUMN_STYLES["info"]["fill_color"],
        fill_type="solid",
    )

    header_fills = {
        "editable": editable_fill,
        "computed": computed_fill,
        "info": info_fill,
    }

    green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    red_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")

    status_fills = {
        "GREEN": green_fill,
        "YELLOW": yellow_fill,
        "RED": red_fill,
    }

    # Write header row with color-coded columns
    for col_idx, (col_name, _, _) in enumerate(ALBUMS_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.border = header_border
        cell.alignment = Alignment(horizontal="center")

        # Apply category-based fill color
        style_category = COLUMN_STYLE_MAP.get(col_name, "info")
        cell.fill = header_fills[style_category]

    # Write album rows
    for row_idx, album in enumerate(albums, start=2):
        # Get existing user data for this album
        existing = existing_data.get(album.album_id, {})

        # Compute default action
        default_action = compute_default_action(album)

        # Prepare row data
        row_data = {
            "album_id": album.album_id,
            "source_path": str(album.source_path.relative_to(library_root)),
            "artist": album.metadata.artist,
            "album": album.metadata.album,
            "year": album.metadata.year,
            "track_count": album.track_count,
            "source_formats": ";".join(sorted(f.value for f in album.source_formats)),
            "max_sr_hz": album.max_sample_rate,
            "max_bit_depth": album.max_bit_depth,
            "default_action": default_action.value,
            "user_action": existing.get("user_action", ""),
            "aac_target_kbps": existing.get("aac_target_kbps", ""),
            "skip": existing.get("skip", ""),
            "tag_status": album.tag_status.value,
            "art_status": album.art_status.value,
            "plan_hash": "",
            "last_built_at": existing.get("last_built_at", ""),
            "error_code": "",
            "notes": "; ".join(album.status_notes) if album.status_notes else "",
        }

        # Preserve user notes if they added any
        if existing.get("notes") and not row_data["notes"]:
            row_data["notes"] = existing["notes"]
        elif existing.get("notes") and row_data["notes"]:
            # Append tool notes to user notes
            row_data["notes"] = f"{existing['notes']}; {row_data['notes']}"

        # Write cells
        for col_idx, (col_name, _, _) in enumerate(ALBUMS_COLUMNS, start=1):
            value = row_data.get(col_name, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)

            # Apply conditional formatting for status columns
            if col_name in ("tag_status", "art_status") and value in status_fills:
                cell.fill = status_fills[value]
                cell.alignment = Alignment(horizontal="center")

            # Highlight error_code if present
            if col_name == "error_code" and value:
                cell.fill = red_fill

    # Set column widths
    column_widths = {
        "album_id": 18,
        "source_path": 50,
        "artist": 25,
        "album": 30,
        "year": 8,
        "track_count": 10,
        "source_formats": 15,
        "max_sr_hz": 12,
        "max_bit_depth": 12,
        "default_action": 15,
        "user_action": 15,
        "aac_target_kbps": 15,
        "skip": 8,
        "tag_status": 12,
        "art_status": 12,
        "plan_hash": 18,
        "last_built_at": 20,
        "error_code": 15,
        "notes": 40,
    }

    for col_idx, (col_name, _, _) in enumerate(ALBUMS_COLUMNS, start=1):
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = column_widths.get(col_name, 15)

    # Freeze header row
    ws.freeze_panes = "A2"

    # Add auto-filter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(ALBUMS_COLUMNS))}{len(albums) + 1}"


def update_xlsx_album(
    xlsx_path: Path,
    album_id: str,
    updates: dict,
) -> None:
    """
    Update a single album row in existing XLSX.

    Args:
        xlsx_path: Path to XLSX file
        album_id: ID of album to update
        updates: Dict of column_name -> new_value
    """
    from openpyxl import load_workbook

    check_xlsx_lock(xlsx_path)

    wb = load_workbook(xlsx_path)
    ws = wb["Albums"]

    # Find album row
    album_id_col = 1  # album_id is first column
    for row_idx in range(2, ws.max_row + 1):
        if ws.cell(row=row_idx, column=album_id_col).value == album_id:
            # Update cells
            for col_name, value in updates.items():
                col_idx = None
                for idx, (name, _, _) in enumerate(ALBUMS_COLUMNS, start=1):
                    if name == col_name:
                        col_idx = idx
                        break
                if col_idx:
                    ws.cell(row=row_idx, column=col_idx, value=value)
            break

    # Save with atomic write
    temp_path = xlsx_path.parent / f"{xlsx_path.name}.tmp"
    try:
        wb.save(temp_path)
        os.replace(temp_path, xlsx_path)
    finally:
        if temp_path.exists():
            temp_path.unlink()


def _write_reference_sheet(wb: Workbook) -> None:
    """Write Reference tab with enum options and column legend."""
    ws = wb.create_sheet("Reference", 2)

    # Styles
    header_font = Font(bold=True, size=12)
    subheader_font = Font(bold=True)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    # Column style fills
    editable_fill = PatternFill(
        start_color=COLUMN_STYLES["editable"]["fill_color"],
        end_color=COLUMN_STYLES["editable"]["fill_color"],
        fill_type="solid",
    )
    computed_fill = PatternFill(
        start_color=COLUMN_STYLES["computed"]["fill_color"],
        end_color=COLUMN_STYLES["computed"]["fill_color"],
        fill_type="solid",
    )
    info_fill = PatternFill(
        start_color=COLUMN_STYLES["info"]["fill_color"],
        end_color=COLUMN_STYLES["info"]["fill_color"],
        fill_type="solid",
    )

    row = 1

    # Column Legend Section
    ws.cell(row=row, column=1, value="COLUMN LEGEND").font = header_font
    ws.cell(row=row, column=1).fill = header_fill
    row += 1

    ws.cell(row=row, column=1, value="Column headers are color-coded to indicate their purpose:")
    row += 2

    # Editable columns
    ws.cell(row=row, column=1, value="GREEN - Editable").font = subheader_font
    ws.cell(row=row, column=1).fill = editable_fill
    ws.cell(row=row, column=2, value="You can modify these values to control conversion")
    row += 1

    editable_cols = [col for col, style in COLUMN_STYLE_MAP.items() if style == "editable"]
    ws.cell(row=row, column=1, value=f"  Columns: {', '.join(editable_cols)}")
    row += 2

    # Computed columns
    ws.cell(row=row, column=1, value="BLUE - Computed").font = subheader_font
    ws.cell(row=row, column=1).fill = computed_fill
    ws.cell(row=row, column=2, value="Tool-generated analysis (do not edit)")
    row += 1

    computed_cols = [col for col, style in COLUMN_STYLE_MAP.items() if style == "computed"]
    ws.cell(row=row, column=1, value=f"  Columns: {', '.join(computed_cols)}")
    row += 2

    # Info columns
    ws.cell(row=row, column=1, value="GRAY - Informational").font = subheader_font
    ws.cell(row=row, column=1).fill = info_fill
    ws.cell(row=row, column=2, value="Source metadata (read-only)")
    row += 1

    info_cols = [col for col, style in COLUMN_STYLE_MAP.items() if style == "info"]
    ws.cell(row=row, column=1, value=f"  Columns: {', '.join(info_cols)}")
    row += 3

    # Action Options Section
    ws.cell(row=row, column=1, value="ACTION OPTIONS (for user_action column)").font = header_font
    ws.cell(row=row, column=1).fill = header_fill
    row += 2

    ws.cell(row=row, column=1, value="Action").font = subheader_font
    ws.cell(row=row, column=2, value="Description").font = subheader_font
    row += 1

    for action, description in ACTION_OPTIONS:
        ws.cell(row=row, column=1, value=action)
        ws.cell(row=row, column=2, value=description)
        row += 1

    row += 2

    # AAC Bitrate Options
    ws.cell(row=row, column=1, value="AAC BITRATE OPTIONS (for aac_target_kbps column)").font = header_font
    ws.cell(row=row, column=1).fill = header_fill
    row += 2

    ws.cell(row=row, column=1, value=f"Allowed values: {', '.join(str(b) for b in AAC_BITRATE_OPTIONS)} kbps")
    ws.cell(row=row, column=2, value="Default: 256 kbps")
    row += 3

    # Status Values Section
    ws.cell(row=row, column=1, value="STATUS VALUES (tag_status / art_status)").font = header_font
    ws.cell(row=row, column=1).fill = header_fill
    row += 2

    green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    red_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")

    status_fills = {"GREEN": green_fill, "YELLOW": yellow_fill, "RED": red_fill}

    for status, description in STATUS_VALUES:
        ws.cell(row=row, column=1, value=status).fill = status_fills[status]
        ws.cell(row=row, column=2, value=description)
        row += 1

    row += 2

    # Skip Column
    ws.cell(row=row, column=1, value="SKIP COLUMN").font = header_font
    ws.cell(row=row, column=1).fill = header_fill
    row += 2

    ws.cell(row=row, column=1, value="TRUE")
    ws.cell(row=row, column=2, value="Skip this album entirely (leave blank or FALSE to include)")
    row += 3

    # Quick Reference
    ws.cell(row=row, column=1, value="QUICK WORKFLOW").font = header_font
    ws.cell(row=row, column=1).fill = header_fill
    row += 2

    workflow_steps = [
        "1. Review the Albums tab - check tag_status and art_status columns",
        "2. For albums with RED status, consider fixing source files or skipping",
        "3. To override default conversion: set user_action to your preferred action",
        "4. To convert to AAC: set user_action=AAC and optionally set aac_target_kbps",
        "5. To skip an album: set skip=TRUE",
        "6. Save the XLSX and run: yangon apply --xlsx <this_file> --out <output_dir>",
    ]

    for step in workflow_steps:
        ws.cell(row=row, column=1, value=step)
        row += 1

    # Set column widths
    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["B"].width = 70
