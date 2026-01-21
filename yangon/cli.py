"""CLI entry point using Click."""

import os
import subprocess
import sys
from pathlib import Path

import click
from rich.console import Console
from rich.table import Table

from yangon import __version__
from yangon.models.config import ApplyConfig, Config, ScanConfig
from yangon.models.status import ArtStatus, TagStatus


console = Console()

# Supported plan file formats
PLAN_FORMATS = ["xlsx", "tsv", "csv"]


def detect_plan_format(path: Path) -> str:
    """Detect plan format from file extension."""
    suffix = path.suffix.lower()
    if suffix == ".tsv":
        return "tsv"
    elif suffix == ".csv":
        return "csv"
    elif suffix in (".xlsx", ".xls"):
        return "xlsx"
    else:
        return "xlsx"  # Default to xlsx


def get_plan_path(base_path: Path, fmt: str) -> Path:
    """Get plan file path with appropriate extension."""
    if fmt == "tsv":
        return base_path.with_suffix(".tsv")
    elif fmt == "csv":
        return base_path.with_suffix(".csv")
    else:
        return base_path.with_suffix(".xlsx")


def check_ffmpeg() -> bool:
    """Check if FFmpeg is available."""
    try:
        result = subprocess.run(
            ["ffmpeg", "-version"],
            capture_output=True,
            timeout=5,
        )
        return result.returncode == 0
    except Exception:
        return False


@click.group()
@click.version_option(version=__version__)
def cli():
    """
    yangon - iPod-compatible audio library converter.

    Scan a music library, generate an XLSX decision sheet,
    then build an iPod-ready output library.
    """
    pass


@cli.command()
@click.option(
    "--library",
    "-l",
    type=click.Path(exists=True, file_okay=False, path_type=Path),
    required=True,
    help="Path to music library root directory",
)
@click.option(
    "--plan",
    "-p",
    "plan_path",
    type=click.Path(path_type=Path),
    required=True,
    help="Path to output plan file (XLSX, TSV, or CSV)",
)
@click.option(
    "--format",
    "-f",
    "plan_format",
    type=click.Choice(PLAN_FORMATS),
    default=None,
    help="Plan file format (auto-detected from extension if not specified)",
)
@click.option(
    "--recreate",
    is_flag=True,
    help="Recreate plan from scratch (discard user edits)",
)
@click.option(
    "--normalize-tags",
    is_flag=True,
    help="Enable tag normalization",
)
@click.option(
    "--threads",
    "-t",
    type=int,
    default=32,
    help="Number of threads for parallel scanning",
)
@click.option(
    "--no-tui",
    is_flag=True,
    help="Disable TUI progress display",
)
@click.option(
    "--compact",
    is_flag=True,
    help="Use compact TUI mode (for small terminals)",
)
def scan(
    library: Path,
    plan_path: Path,
    plan_format: str | None,
    recreate: bool,
    normalize_tags: bool,
    threads: int,
    no_tui: bool,
    compact: bool,
):
    """
    Scan music library and generate/update plan file.

    Walks the library directory, analyzes each album's tracks,
    and creates a plan file (XLSX, TSV, or CSV) with conversion options.

    Examples:
        yangon scan -l /music -p plan.xlsx       # XLSX format (Excel)
        yangon scan -l /music -p plan.tsv        # TSV format (text editor)
        yangon scan -l /music -p plan.csv        # CSV format (text editor)
    """
    if not check_ffmpeg():
        console.print("[red]Error: FFmpeg not found. Please install FFmpeg.[/red]")
        sys.exit(1)

    library = library.resolve()
    plan_path = plan_path.resolve()

    # Determine format
    fmt = plan_format or detect_plan_format(plan_path)
    if fmt == "tsv" and plan_path.suffix.lower() != ".tsv":
        plan_path = plan_path.with_suffix(".tsv")
    elif fmt == "csv" and plan_path.suffix.lower() != ".csv":
        plan_path = plan_path.with_suffix(".csv")
    elif fmt == "xlsx" and plan_path.suffix.lower() not in (".xlsx", ".xls"):
        plan_path = plan_path.with_suffix(".xlsx")

    config = ScanConfig(
        library_root=library,
        xlsx_path=plan_path,  # Used for both formats
        recreate=recreate,
        normalize_tags=normalize_tags,
        threads=threads,
        show_tui=not no_tui,
    )

    console.print(f"[blue]Scanning library:[/blue] {library}")
    console.print(f"[blue]Output plan ({fmt.upper()}):[/blue] {plan_path}")

    from yangon.scanner.detector import scan_library
    from yangon.tui.dashboard import run_with_dashboard
    from yangon.tui.events import (
        EventBus,
        ScanCompleteEvent,
        ScanProgressEvent,
        ScanStartEvent,
    )

    event_bus = EventBus()
    albums = []

    def do_scan():
        nonlocal albums

        def progress_callback(current: int, total: int, name: str):
            if current == 0:
                event_bus.emit(ScanStartEvent(total_dirs=total))
            event_bus.emit(ScanProgressEvent(
                current=current,
                total=total,
                current_dir=name,
            ))

        albums = scan_library(config, progress_callback=progress_callback)

        total_tracks = sum(a.track_count for a in albums)
        event_bus.emit(ScanCompleteEvent(
            albums_found=len(albums),
            tracks_found=total_tracks,
        ))

    run_with_dashboard(event_bus, do_scan, show_tui=not no_tui, compact=compact)

    # Write plan file
    if albums:
        console.print(f"\n[blue]Writing {fmt.upper()} plan...[/blue]")

        if fmt in ("tsv", "csv"):
            from yangon.csv_io.writer import write_csv_plan
            use_tsv = fmt == "tsv"
            write_csv_plan(albums, plan_path, library, preserve_user_edits=not recreate, use_tsv=use_tsv)
        else:
            from yangon.xlsx.writer import write_xlsx
            write_xlsx(albums, plan_path, library, preserve_user_edits=not recreate)

        console.print(f"[green]Done![/green] Found {len(albums)} albums")
        console.print(f"Plan saved to: {plan_path}")
    else:
        console.print("[yellow]No albums found in library.[/yellow]")


@cli.command()
@click.option(
    "--plan",
    "-p",
    "plan_path",
    type=click.Path(exists=True, path_type=Path),
    required=True,
    help="Path to plan file (XLSX, TSV, or CSV)",
)
@click.option(
    "--out",
    "-o",
    type=click.Path(path_type=Path),
    required=True,
    help="Output directory for converted library",
)
@click.option(
    "--dry-run",
    is_flag=True,
    help="Show what would be done without converting",
)
@click.option(
    "--fail-fast",
    is_flag=True,
    help="Stop on first error",
)
@click.option(
    "--force",
    is_flag=True,
    help="Rebuild all tracks (ignore cache)",
)
@click.option(
    "--threads",
    "-t",
    type=int,
    default=None,
    help="Number of threads for parallel conversion (default: CPU count)",
)
@click.option(
    "--no-tui",
    is_flag=True,
    help="Disable TUI progress display",
)
@click.option(
    "--compact",
    is_flag=True,
    help="Use compact TUI mode (for small terminals)",
)
def apply(
    plan_path: Path,
    out: Path,
    dry_run: bool,
    fail_fast: bool,
    force: bool,
    threads: int | None,
    no_tui: bool,
    compact: bool,
):
    """
    Apply plan decisions to build output library.

    Reads the plan file (XLSX, TSV, or CSV), resolves actions for each album,
    and converts/copies tracks to the output directory.

    Examples:
        yangon apply --plan plan.xlsx --out /output
        yangon apply --plan plan.tsv --out /output
    """
    if not check_ffmpeg():
        console.print("[red]Error: FFmpeg not found. Please install FFmpeg.[/red]")
        sys.exit(1)

    plan_path = plan_path.resolve()
    out = out.resolve()

    if threads is None:
        threads = os.cpu_count() or 8

    # Detect format
    fmt = detect_plan_format(plan_path)

    config = ApplyConfig(
        xlsx_path=plan_path,  # Used for both formats
        output_root=out,
        dry_run=dry_run,
        fail_fast=fail_fast,
        force=force,
        threads=threads,
        show_tui=not no_tui,
    )
    global_config = Config()

    console.print(f"[blue]Reading plan ({fmt.upper()}):[/blue] {plan_path}")
    console.print(f"[blue]Output directory:[/blue] {out}")
    if dry_run:
        console.print("[yellow]DRY RUN - no files will be written[/yellow]")

    from yangon.converter.pipeline import (
        ConversionPipeline,
        JobCompletedEvent,
        JobErrorEvent,
        JobStartedEvent,
    )
    from yangon.planner.resolver import resolve_build_plan
    from yangon.scanner.detector import scan_library
    from yangon.tui.dashboard import run_with_dashboard
    from yangon.tui.events import (
        BuildCompleteEvent,
        BuildProgressEvent,
        BuildStartEvent,
        EventBus,
        TrackCompleteEvent,
        TrackErrorEvent,
        TrackStartEvent,
    )

    # Read decisions based on format
    if fmt in ("tsv", "csv"):
        from yangon.csv_io.reader import get_csv_decisions, get_csv_library_root
        decisions_list = get_csv_decisions(plan_path)
        library_root = get_csv_library_root(plan_path)
    else:
        from yangon.xlsx.reader import get_album_decisions
        decisions_list = get_album_decisions(plan_path)
        # Get library root from XLSX
        from openpyxl import load_workbook
        wb = load_workbook(plan_path, read_only=True)
        library_root = None
        if "Summary" in wb.sheetnames:
            for row in wb["Summary"].iter_rows(min_row=1, max_row=10, max_col=2, values_only=True):
                if row[0] == "library_root":
                    library_root = Path(row[1])
                    break
        wb.close()

    decisions = {d["album_id"]: d for d in decisions_list}

    if not library_root or not library_root.exists():
        console.print(f"[red]Error: Library root not found: {library_root}[/red]")
        sys.exit(1)

    console.print(f"[blue]Library root:[/blue] {library_root}")

    # Re-scan to get album objects
    scan_config = ScanConfig(library_root=library_root, xlsx_path=plan_path)
    albums = scan_library(scan_config)

    # Resolve build plan
    plan = resolve_build_plan(albums, decisions, config, global_config.tool_version)

    console.print(f"[blue]Jobs to process:[/blue] {plan.total_tracks}")
    console.print(f"[blue]Albums skipped:[/blue] {len(plan.skipped_albums)}")

    if plan.validation_errors:
        console.print(f"[yellow]Validation errors:[/yellow] {len(plan.validation_errors)}")
        for err in plan.validation_errors[:5]:
            console.print(f"  {err['album_id']}: {err['message']}")

    if not plan.jobs:
        console.print("[yellow]No jobs to process.[/yellow]")
        return

    event_bus = EventBus()
    results = []

    # Build album lookup for display names
    album_lookup = {a.album_id: a for a in albums}

    def event_adapter(pipeline_event):
        """Adapt pipeline events to TUI events."""
        if isinstance(pipeline_event, JobStartedEvent):
            job = pipeline_event.job
            if not job:
                return

            # Get album display name
            album = album_lookup.get(job.album_id)
            album_name = ""
            if album and album.metadata:
                artist = album.metadata.album_artist or album.metadata.artist or "Unknown"
                title = album.metadata.album or "Unknown"
                album_name = f"{artist} / {title}"

            event_bus.emit(TrackStartEvent(
                album_id=job.album_id,
                track_path=str(job.source_path),
                action=job.action.value,
            ))
            event_bus.emit(BuildProgressEvent(
                completed=pipeline.stats.completed_jobs,
                failed=pipeline.stats.failed_jobs,
                cached=pipeline.stats.cached_jobs,
                total=pipeline.stats.total_jobs,
                current_album=album_name,
                current_track=job.source_path.name,
            ))
        elif isinstance(pipeline_event, JobCompletedEvent):
            job = pipeline_event.job
            result = pipeline_event.result
            if not job:
                return
            event_bus.emit(TrackCompleteEvent(
                album_id=job.album_id,
                track_path=str(job.source_path),
                output_path=str(result.output_path) if result and result.output_path else "",
                success=result.success if result else False,
            ))
        elif isinstance(pipeline_event, JobErrorEvent):
            job = pipeline_event.job
            if not job:
                return
            event_bus.emit(TrackErrorEvent(
                album_id=job.album_id,
                track_path=str(job.source_path),
                error_code="ENCODE_FAIL",
                error_message=pipeline_event.error,
            ))

    pipeline = ConversionPipeline(
        config=config,
        global_config=global_config,
        event_callback=event_adapter,
    )

    def do_apply():
        nonlocal results
        event_bus.emit(BuildStartEvent(total_jobs=plan.total_tracks))
        results = pipeline.execute(plan, dry_run=dry_run)
        event_bus.emit(BuildCompleteEvent(
            total=len(results),
            succeeded=sum(1 for r in results if r.success),
            failed=sum(1 for r in results if not r.success),
            cached=pipeline.stats.cached_jobs,
        ))

    run_with_dashboard(event_bus, do_apply, show_tui=not no_tui, compact=compact)

    # Summary
    succeeded = sum(1 for r in results if r.success)
    failed = sum(1 for r in results if not r.success)

    console.print(f"\n[green]Complete![/green]")
    console.print(f"  Succeeded: {succeeded}")
    console.print(f"  Failed: {failed}")
    console.print(f"  Cached: {pipeline.stats.cached_jobs}")

    if failed > 0:
        sys.exit(1)


@cli.command()
@click.option(
    "--plan",
    "-p",
    "plan_path",
    type=click.Path(exists=True, path_type=Path),
    required=True,
    help="Path to plan file (XLSX, TSV, or CSV)",
)
def status(plan_path: Path):
    """
    Display summary from plan file.

    Shows album counts, status distribution, and action summary.
    Works with XLSX, TSV, and CSV formats.
    """
    plan_path = plan_path.resolve()
    fmt = detect_plan_format(plan_path)

    console.print(f"\n[bold]Plan Summary ({fmt.upper()}): {plan_path.name}[/bold]\n")

    tag_counts: dict[str, int] = {}
    art_counts: dict[str, int] = {}
    action_counts: dict[str, int] = {}

    if fmt in ("tsv", "csv"):
        # CSV/TSV format
        from yangon.csv_io.reader import get_csv_summary
        summary = get_csv_summary(plan_path)

        summary_table = Table(show_header=False, box=None)
        summary_table.add_column("Key", style="bold")
        summary_table.add_column("Value")

        summary_table.add_row("library_root", str(summary.get("library_root", "Unknown")))
        summary_table.add_row("total_albums", str(summary.get("total_albums", 0)))
        summary_table.add_row("total_tracks", str(summary.get("total_tracks", 0)))
        summary_table.add_row("total_size_mb", f'{summary.get("total_size_mb", 0):.1f}')
        summary_table.add_row("created_at", str(summary.get("created_at", "Unknown")))

        console.print(summary_table)

        tag_counts = summary.get("tag_status_counts", {})
        art_counts = summary.get("art_status_counts", {})
        action_counts = summary.get("action_counts", {})
    else:
        # XLSX format
        from openpyxl import load_workbook

        wb = load_workbook(plan_path, read_only=True)

        if "Summary" in wb.sheetnames:
            summary_table = Table(show_header=False, box=None)
            summary_table.add_column("Key", style="bold")
            summary_table.add_column("Value")

            for row in wb["Summary"].iter_rows(values_only=True):
                if row[0] and row[1]:
                    summary_table.add_row(str(row[0]), str(row[1]))

            console.print(summary_table)

        # Count albums by status from XLSX
        if "Albums" in wb.sheetnames:
            ws = wb["Albums"]
            header = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]

            col_map = {h: i for i, h in enumerate(header) if h}

            tag_idx = col_map.get("tag_status")
            art_idx = col_map.get("art_status")
            action_idx = col_map.get("default_action")
            user_action_idx = col_map.get("user_action")

            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or not row[0]:
                    continue

                if tag_idx is not None and row[tag_idx]:
                    tag_counts[row[tag_idx]] = tag_counts.get(row[tag_idx], 0) + 1
                if art_idx is not None and row[art_idx]:
                    art_counts[row[art_idx]] = art_counts.get(row[art_idx], 0) + 1

                action = row[user_action_idx] if user_action_idx and row[user_action_idx] else row[action_idx]
                if action:
                    action_counts[action] = action_counts.get(action, 0) + 1

        wb.close()

    # Display status tables
    if tag_counts:
        console.print("\n[bold]Tag Status:[/bold]")
        tag_table = Table(show_header=True)
        tag_table.add_column("Status")
        tag_table.add_column("Count", justify="right")
        for stat in ["GREEN", "YELLOW", "RED"]:
            count = tag_counts.get(stat, 0)
            style = {"GREEN": "green", "YELLOW": "yellow", "RED": "red"}.get(stat, "")
            tag_table.add_row(stat, str(count), style=style)
        console.print(tag_table)

    if art_counts:
        console.print("\n[bold]Art Status:[/bold]")
        art_table = Table(show_header=True)
        art_table.add_column("Status")
        art_table.add_column("Count", justify="right")
        for stat in ["GREEN", "YELLOW", "RED"]:
            count = art_counts.get(stat, 0)
            style = {"GREEN": "green", "YELLOW": "yellow", "RED": "red"}.get(stat, "")
            art_table.add_row(stat, str(count), style=style)
        console.print(art_table)

    if action_counts:
        console.print("\n[bold]Actions:[/bold]")
        action_table = Table(show_header=True)
        action_table.add_column("Action")
        action_table.add_column("Count", justify="right")
        for action, count in sorted(action_counts.items()):
            action_table.add_row(action, str(count))
        console.print(action_table)


if __name__ == "__main__":
    cli()
