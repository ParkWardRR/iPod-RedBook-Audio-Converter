"""XLSX reading and parsing."""

from pathlib import Path

from openpyxl import load_workbook

from ipodrb.models.plan import Action
from ipodrb.xlsx.schemas import ALBUMS_COLUMNS, COLUMN_INDEX, SCHEMA_VERSION


class XLSXSchemaError(Exception):
    """XLSX schema version mismatch or invalid structure."""

    pass


def read_xlsx(xlsx_path: Path) -> dict[str, dict]:
    """
    Read XLSX and extract album data keyed by album_id.

    Args:
        xlsx_path: Path to XLSX file

    Returns:
        Dict mapping album_id to row data dict

    Raises:
        XLSXSchemaError: If schema version is incompatible
    """
    xlsx_path = Path(xlsx_path)
    if not xlsx_path.exists():
        return {}

    wb = load_workbook(xlsx_path, read_only=True, data_only=True)

    # Validate schema version from Summary sheet
    if "Summary" in wb.sheetnames:
        summary = wb["Summary"]
        for row in summary.iter_rows(min_row=1, max_row=10, max_col=2, values_only=True):
            if row[0] == "schema_version":
                version = str(row[1]) if row[1] else ""
                if not version.startswith(SCHEMA_VERSION.split(".")[0]):
                    raise XLSXSchemaError(
                        f"XLSX schema version {version} is not compatible with "
                        f"tool version {SCHEMA_VERSION}. Please recreate the sheet."
                    )
                break

    # Read Albums sheet
    if "Albums" not in wb.sheetnames:
        return {}

    ws = wb["Albums"]

    # Validate header row
    header_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    expected_cols = [col[0] for col in ALBUMS_COLUMNS]

    # Build column mapping (handle potential column reordering)
    col_map = {}
    for col_idx, header in enumerate(header_row):
        if header in expected_cols:
            col_map[header] = col_idx

    # Check required columns exist
    required = {"album_id", "user_action", "aac_target_kbps", "skip", "last_built_at", "notes"}
    missing = required - set(col_map.keys())
    if missing:
        raise XLSXSchemaError(f"Missing required columns: {missing}")

    # Read album rows
    result = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[col_map.get("album_id", 0)]:
            continue

        album_id = str(row[col_map["album_id"]])

        # Extract user-editable and relevant columns
        row_data = {}
        for col_name in col_map:
            value = row[col_map[col_name]]
            if value is not None:
                row_data[col_name] = value

        # Normalize user_action
        if "user_action" in row_data:
            user_action = row_data["user_action"]
            if user_action:
                # Validate it's a valid action
                try:
                    Action(user_action)
                except ValueError:
                    row_data["user_action_error"] = f"Invalid action: {user_action}"
                    row_data["user_action"] = ""

        # Normalize skip to boolean string
        if "skip" in row_data:
            skip_val = row_data["skip"]
            if isinstance(skip_val, bool):
                row_data["skip"] = "TRUE" if skip_val else ""
            elif isinstance(skip_val, str):
                row_data["skip"] = "TRUE" if skip_val.upper() in ("TRUE", "YES", "1") else ""

        result[album_id] = row_data

    wb.close()
    return result


def get_album_decisions(xlsx_path: Path) -> list[dict]:
    """
    Get album decisions from XLSX for apply command.

    Returns list of dicts with:
        - album_id
        - source_path
        - user_action (or None if using default)
        - aac_target_kbps (or None)
        - skip (bool)
    """
    xlsx_path = Path(xlsx_path)
    if not xlsx_path.exists():
        raise FileNotFoundError(f"XLSX not found: {xlsx_path}")

    wb = load_workbook(xlsx_path, read_only=True, data_only=True)

    if "Albums" not in wb.sheetnames:
        return []

    ws = wb["Albums"]

    # Build column mapping
    header_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    col_map = {}
    for col_idx, header in enumerate(header_row):
        col_map[header] = col_idx

    decisions = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[col_map.get("album_id", 0)]:
            continue

        album_id = str(row[col_map["album_id"]])
        source_path = row[col_map.get("source_path", 1)] or ""
        default_action = row[col_map.get("default_action", 9)] or ""
        user_action = row[col_map.get("user_action", 10)] or ""
        aac_kbps = row[col_map.get("aac_target_kbps", 11)]
        skip_val = row[col_map.get("skip", 12)]

        # Determine resolved action
        resolved_action = user_action if user_action else default_action

        # Parse skip
        skip = False
        if skip_val:
            if isinstance(skip_val, bool):
                skip = skip_val
            elif isinstance(skip_val, str):
                skip = skip_val.upper() in ("TRUE", "YES", "1")

        decisions.append({
            "album_id": album_id,
            "source_path": source_path,
            "default_action": default_action,
            "user_action": user_action or None,
            "resolved_action": resolved_action,
            "aac_target_kbps": int(aac_kbps) if aac_kbps else None,
            "skip": skip,
        })

    wb.close()
    return decisions
